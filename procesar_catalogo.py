"""
Procesador completo del catálogo de Bicicmarket
================================================
Hace todo el flujo automáticamente:

1. (Opcional) Descarga el catálogo desde la URL de asociado
2. Filtra solo Bicicletas + Bicicletas Eléctricas (renombra esta última)
3. Añade columnas con fórmulas: PCA+IVA+ENVÍO, MARGEN, % DTO., DTO. SOBRE PVPR,
   PRECIO WEB, BENEFICIO (las del primo)
4. Scrapeea el PRECIO_BICIMARKET real de la web
5. Calcula PRECIO_TIENDA con los descuentos (-50€ / -90€ para marcas especiales)

Uso:
    # Descargar y procesar todo de una vez:
    python procesar_catalogo.py --descargar

    # Procesar un fichero ya descargado:
    python procesar_catalogo.py reporte-catalogo.xlsx

    # Modo prueba con pocos productos:
    python procesar_catalogo.py --descargar --limite 10
"""

import re
import time
import shutil
import argparse
import unicodedata
from pathlib import Path
from datetime import datetime

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Configuración ─────────────────────────────────────────────────────────────

URL_CATALOGO = "https://privado.bicimarket.com/reporting/catalogo_asociados/3fj38gn6nc6q9gnbhf"

CATEGORIA_SLUG = {
    "Bicicletas":             "bicicletas",
    "Bicicletas Eléctricas":  "bicicleta-electrica",
    "Bicicleta Electrica":    "bicicleta-electrica",
}

SUBCATEGORIA_SLUG = {
    "Gravel": "gravel", "MTB 29\"": "mtb-29", "MTB 27.5\"": "mtb-27-5",
    "MTB 26\"": "mtb-26", "Carretera": "carretera", "Urbanas": "urbanas",
    "Hibridas/Trekking": "hibridas-trekking", "Trekking": "hibridas-trekking",
    "Infantiles/Junior": "infantiles-junior", "Plegables": "plegables",
    "BMX": "bmx", "Ciclocross": "ciclocross", "Triatlon": "triatlon",
    "Bicicletas Mujer": "bicicletas-mujer", "Cargo Bike": "cargo-bike",
    "eCargo Bike": "ecargo-bike", "Estáticas/Spinning": "estaticas-spinning",
    "Fat eBike": "fat-ebike", "eBike MTB 29\"": "ebike-mtb-29",
    "eBike MTB 27.5\"": "ebike-mtb-27-5", "eBike Gravel": "ebike-gravel",
    "eBike Carretera": "ebike-carretera",
    "eBike Híbrida/Trekking": "ebike-hibrida-trekking",
    "eBike Urbana": "ebike-urbana", "eBike Plegable": "ebike-plegable",
}

HEADERS_WEB = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept-Language": "es-ES,es;q=0.9",
}

CATEGORIAS_INCLUIR = {"Bicicletas", "Bicicleta Electrica"}
MARCAS_DESCUENTO_ESPECIAL = {"KTM", "Raymon", "Wilier", "Fantic"}
DESCUENTO_ESPECIAL = 90
DESCUENTO_NORMAL = 50

# ── Funciones de utilidad ─────────────────────────────────────────────────────

def slugify(text):
    text = unicodedata.normalize("NFKD", str(text))
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def build_url(id_producto, marca, producto, categoria, subcategoria):
    cat = CATEGORIA_SLUG.get(categoria, slugify(categoria))
    sub = SUBCATEGORIA_SLUG.get(subcategoria, slugify(subcategoria))
    nombre = slugify(f"{marca} {producto}")
    return f"https://www.bicimarket.com/es/{cat}/{sub}/{id_producto}-{nombre}.html"


def get_precio(url, session, retries=3):
    for _ in range(retries):
        try:
            r = session.get(url, headers=HEADERS_WEB, timeout=15)
            if r.status_code == 404:
                return None
            if r.status_code != 200:
                time.sleep(2)
                continue
            soup = BeautifulSoup(r.text, "lxml")
            el = soup.select_one("[itemprop='price']")
            if el:
                raw = el.get("content") or el.get_text()
                raw = re.sub(r"[€$£\s]", "", str(raw))
                if re.search(r"\d\.\d{3},", raw):
                    raw = raw.replace(".", "").replace(",", ".")
                elif "," in raw and "." not in raw:
                    raw = raw.replace(",", ".")
                raw = re.sub(r"[^\d.]", "", raw)
                try:
                    return float(raw)
                except ValueError:
                    return None
            return None
        except requests.RequestException:
            time.sleep(3)
    return None


def descargar_catalogo(destino):
    print(f"⬇️  Descargando catálogo de Bicicmarket...")
    with requests.get(URL_CATALOGO, headers=HEADERS_WEB, timeout=300, stream=True) as r:
        r.raise_for_status()
        total = 0
        with open(destino, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                f.write(chunk)
                total += len(chunk)
                print(f"   {total/1024:.0f} KB descargados...", end="\r")
    print(f"\n   ✅ Guardado en: {destino} ({total/1024:.0f} KB)")


def calcular_precio_tienda(marca, precio_bicimarket, coste_iva_envio):
    """
    Aplica el descuento según la marca:
    - KTM, Raymon, Wilier, Fantic: -90€
    - Resto: -50€
    Si el resultado es menor que el coste, deja el precio Bicimarket sin tocar.
    """
    if precio_bicimarket is None:
        return None
    marca_norm = str(marca).strip()
    descuento = DESCUENTO_ESPECIAL if marca_norm in MARCAS_DESCUENTO_ESPECIAL else DESCUENTO_NORMAL
    precio_tienda = precio_bicimarket - descuento
    if coste_iva_envio is not None and precio_tienda < coste_iva_envio:
        return precio_bicimarket
    return precio_tienda


# ── Procesador principal ──────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("excel", nargs="?", default=None,
                        help="Fichero de entrada (omitir si usas --descargar)")
    parser.add_argument("--descargar", action="store_true",
                        help="Descargar el catálogo de Bicicmarket antes de procesarlo")
    parser.add_argument("--limite", type=int, default=None,
                        help="Limitar a N productos (para pruebas)")
    parser.add_argument("--delay", type=float, default=1.0,
                        help="Segundos entre peticiones (default: 1.0)")
    parser.add_argument("--output", default=None,
                        help="Nombre del fichero de salida")
    args = parser.parse_args()

    # ── 1. Obtener fichero de entrada ─────────────────────────────────────────
    if args.descargar:
        fecha = datetime.now().strftime("%Y%m%d")
        entrada = f"reporte-catalogo-asociados_{fecha}.xlsx"
        descargar_catalogo(entrada)
    elif args.excel:
        entrada = args.excel
    else:
        parser.error("Debes indicar un fichero o usar --descargar")

    output = args.output or entrada.replace(".xlsx", "_PROCESADO.xlsx")

    # ── 2. Leer catálogo original ─────────────────────────────────────────────
    print(f"\n📂 Leyendo: {entrada}")
    df = pd.read_excel(entrada)
    df.columns = df.columns.str.strip()
    print(f"   {len(df):,} filas en el catálogo original")

    # ── 3. Filtrar solo bicicletas y renombrar categoría ──────────────────────
    df = df[df["Categoria"].isin(CATEGORIAS_INCLUIR)].copy()
    df["Categoria"] = df["Categoria"].replace({"Bicicleta Electrica": "Bicicletas Eléctricas"})
    print(f"   {len(df):,} filas tras filtrar bicicletas | "
          f"{df['ID Producto'].nunique():,} productos únicos")

    # ── 4. Scrapeear precios web ──────────────────────────────────────────────
    productos_unicos = df.drop_duplicates("ID Producto")[
        ["ID Producto", "Marca", "Producto", "Categoria", "Subcategoria"]
    ].reset_index(drop=True)

    if args.limite:
        productos_unicos = productos_unicos.head(args.limite)
        # Filtrar el df principal también para que coincida
        ids_limitados = set(productos_unicos["ID Producto"])
        df = df[df["ID Producto"].isin(ids_limitados)].copy()
        print(f"   ⚠️  Modo prueba: solo {args.limite} productos")

    total = len(productos_unicos)
    print(f"\n🔍 Scrapeando {total} productos (delay {args.delay}s)...\n")

    cache = {}
    session = requests.Session()

    for i, row in productos_unicos.iterrows():
        pid = int(row["ID Producto"])
        url = build_url(pid, row["Marca"], row["Producto"],
                        row["Categoria"], row["Subcategoria"])
        precio = get_precio(url, session)
        cache[pid] = precio
        status = f"✅ {precio:.0f}€" if precio else "❌"
        print(f"   [{i+1:>4}/{total}] {row['Marca']} {row['Producto']} → {status}")
        time.sleep(args.delay)

    # ── 5. Construir DataFrame final con todas las columnas ───────────────────
    precio_bicimarket_por_fila = df["ID Producto"].map(lambda pid: cache.get(int(pid))).tolist()

    # Calcular PCA+IVA+ENVÍO en python para usarlo en PRECIO_TIENDA
    coste_iva_envio = (df["PCA"] * 1.21 + 40).round(2).tolist()

    # PRECIO_TIENDA con la lógica de descuentos
    marcas = df["Marca"].tolist()
    precio_tienda_por_fila = [
        calcular_precio_tienda(m, pb, ce)
        for m, pb, ce in zip(marcas, precio_bicimarket_por_fila, coste_iva_envio)
    ]

    # ── 6. Escribir Excel con formato ─────────────────────────────────────────
    print(f"\n💾 Generando: {output}")

    # Guardar primero los datos básicos
    df.to_excel(output, index=False, engine="openpyxl")

    # Reabrir para añadir columnas con fórmulas y formato
    wb = load_workbook(output)
    ws = wb.active

    # Posición actual de columnas:
    # A: ID Producto, B: ID Variante, C: EAN, D: Categoria, E: Subcategoria,
    # F: Marca, G: Producto, H: Variante, I: Stock, J: PCA, K: PVPR,
    # L: URL Descripcion, M: URL Imagen

    # Vamos a reordenar insertando columnas nuevas entre PVPR (K) y URL Descripcion (L)
    # Nuevo orden: ... K=PVPR | L=PCA+IVA+ENVÍO | M=MARGEN | N=% DTO. | O=DTO. SOBRE PVPR
    #              | P=PRECIO WEB | Q=BENEFICIO | R=PRECIO_BICIMARKET | S=PRECIO_TIENDA
    #              | T=URL Descripcion | U=URL Imagen

    # Mover las dos columnas de URL al final insertando 8 columnas
    # Insertamos 8 columnas en posición L (12)
    for _ in range(8):
        ws.insert_cols(12)

    # Cabeceras nuevas
    cabeceras_nuevas = [
        (12, "PCA+IVA+ENVÍO"),
        (13, "MARGEN"),
        (14, "% DTO."),
        (15, "DTO. SOBRE PVPR"),
        (16, "PRECIO WEB"),
        (17, "BENEFICIO"),
        (18, "PRECIO_BICIMARKET"),
        (19, "PRECIO_TIENDA"),
    ]
    for col, nombre in cabeceras_nuevas:
        c = ws.cell(row=1, column=col, value=nombre)
        c.font = Font(name="Calibri", size=11, bold=True)

    # Estilos de relleno
    fill_pca_iva = PatternFill("solid", start_color="E6B9B8")  # rojo claro
    fill_pct_dto = PatternFill("solid", start_color="FFFF00")  # amarillo
    fill_precio_web = PatternFill("solid", start_color="FFC000")  # naranja
    fill_beneficio = PatternFill("solid", start_color="D7E4BD")  # verde claro
    fill_bicimarket = PatternFill("solid", start_color="FFC000")  # naranja
    fill_tienda = PatternFill("solid", start_color="92D050")  # verde

    n_filas = ws.max_row
    for fila in range(2, n_filas + 1):
        # L: PCA+IVA+ENVÍO = J*1.21+40
        c = ws.cell(row=fila, column=12, value=f"=J{fila}*1.21+40")
        c.fill = fill_pca_iva

        # M: MARGEN = K - L
        ws.cell(row=fila, column=13, value=f"=K{fila}-L{fila}")

        # N: % DTO. (vacío para rellenar a mano)
        c = ws.cell(row=fila, column=14)
        c.fill = fill_pct_dto

        # O: DTO. SOBRE PVPR = K * (N/100)
        ws.cell(row=fila, column=15, value=f"=K{fila}*(N{fila}/100)")

        # P: PRECIO WEB = K - O
        c = ws.cell(row=fila, column=16, value=f"=K{fila}-O{fila}")
        c.fill = fill_precio_web
        c.number_format = "0"

        # Q: BENEFICIO = P - L
        c = ws.cell(row=fila, column=17, value=f"=P{fila}-L{fila}")
        c.fill = fill_beneficio

        # R: PRECIO_BICIMARKET (valor scrapeado)
        idx = fila - 2
        if idx < len(precio_bicimarket_por_fila):
            c = ws.cell(row=fila, column=18, value=precio_bicimarket_por_fila[idx])
            c.fill = fill_bicimarket
            c.number_format = "0"

        # S: PRECIO_TIENDA (valor calculado con descuentos)
        if idx < len(precio_tienda_por_fila):
            c = ws.cell(row=fila, column=19, value=precio_tienda_por_fila[idx])
            c.fill = fill_tienda
            c.number_format = "0"

    # Formato de cabecera (fila 1) — solo Calibri 11 bold como el resto
    for cell in ws[1]:
        if cell.value:
            cell.font = Font(name="Calibri", size=11, bold=True)

    wb.save(output)

    # ── 7. Resumen ────────────────────────────────────────────────────────────
    ok = sum(1 for v in cache.values() if v is not None)
    print(f"\n{'='*55}")
    print(f"✅ Listo")
    print(f"   Productos procesados:    {total}")
    print(f"   Precios web encontrados: {ok}/{total} ({ok/total*100:.0f}%)")
    print(f"   Fichero guardado:        {output}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()